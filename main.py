import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import requests
import threading
import time



# Функция для сохранения данных в таблицу Excel
def save_to_excel(text, photo_link):
    wb = load_workbook('site\data.xlsx')
    ws = wb.active
    # Определение строки для записи
    row_number = ws.max_row + 1
    ws.cell(row=row_number, column=1).value = text
    ws.cell(row=row_number, column=2).value = photo_link+'.jpg'
    wb.save('site\data.xlsx')


def download_file(url, path):
    response = requests.get(url)
    with open(path, "wb") as f:
        f.write(response.content)

def get_photo_link(photo):
    max_size = 0
    max_size_url = ''
    for size in photo['sizes']:
        if size['type'] in ['z', 'w'] and size['height'] * size['width'] > max_size:
            max_size = size['height'] * size['width']
            max_size_url = size['url']
    return max_size_url
        
def download_photo(text, photo_link, photo, path):
    url = get_photo_link(photo)
    if url:
        filename = os.path.basename(path)
        current_dir = os.path.dirname(path)
        extension = os.path.splitext(filename)[1]
        photo_path = os.path.join(current_dir, 'site\images', f'{photo_link}{extension}')
        download_file(url, photo_path)
        save_to_excel(text, photo_link)


# Функция для отправки сообщения
def send_message(user_id, message):
    vk.messages.send(
        user_id=user_id,
        message=message,
        random_id=vk_api.utils.get_random_id(),
    )

# Функция для обработки команды !рек
def handle_command(user_id, command):
    try:
        index = int(command)
        df = pd.read_excel('site\data.xlsx')
        if not df.empty and index <= len(df):
            text = df.iloc[index - 1]['text']
            photo_link = df.iloc[index - 1]['image']
            send_message(user_id, f"Текст рекламы: {text}\nСсылка на картинку: {photo_link}")
        else:
            send_message(user_id, "Указанная строка не найдена.")
    except ValueError:
        send_message(user_id, "Необходимо указать номер строки после команды !рек.")



def publ_post():
    while True:
        # Чтение данных из файла Excel
        df = pd.read_excel('site\\new_data.xlsx')
        if not df.empty:
            # Получение данных из первой строки
            text = df.iloc[0]['Text']
            photo_path = "site\images\\" + df.iloc[0]['Image']
            
            # Удаление первой строки из DataFrame
            df = df.drop(0).reset_index(drop=True)
            # Сохранение изменений обратно в файл Excel
            df.to_excel('site\\new_data.xlsx', index=False)

            try:
                # Получаем URL для загрузки фото
                upload_url = vk_session_post.method('photos.getWallUploadServer')['upload_url']
                
                # Загружаем фото на сервер
                with open(photo_path, 'rb') as file:
                    response = requests.post(upload_url, files={'photo': file}).json()

                # Сохраняем фото на сервере ВКонтакте
                photo = vk_session_post.method('photos.saveWallPhoto', {
                    'photo': response['photo'],
                    'server': response['server'],
                    'hash': response['hash']
                })[0]

                photo_attachment = f"photo{photo['owner_id']}_{photo['id']}"

                # Опубликование поста с фото
                vk_session_post.method("wall.post", {
                    "owner_id": id_group,
                    "from_group": 1,
                    "message": text,
                    "attachments": photo_attachment

                })
                vk_post.wall.post
                print(f"Пост опубликован: Текст - {text}, Фото - {photo_path}")
            except Exception as e:
                print(f"Ошибка при загрузке фото или публикации поста: {e}")
        else:
            print("Файл Excel пустой. Повторная попытка через 10 минут.")
            time.sleep(60)  # Задержка выполнения на 10 минут (600 секунд)



# Функция для обработки рекламы
def handle_advertisement(user_id, event):
    # Отправляем сообщение с вопросами пользователю
    send_message(user_id, message="Для сохранения объявления, пожалуйста, ответьте на вопросы.")

    # Ожидаем ответа пользователя
    organization_name = "\n" + wait_for_user_response(user_id, "Введите название организации:")
    website_link = "\n" + wait_for_user_response(user_id, "Введите ссылку на сайт:")
    job_title = "\n" + wait_for_user_response(user_id, "Введите название вакансии:")
    job_function = "\n" + wait_for_user_response(user_id, "Введите функционал:")
    job_requirements = "\n" + wait_for_user_response(user_id, "Введите требования к соискателю:")
    job_conditions = "\n" + wait_for_user_response(user_id, "Введите условия работы:")
    contact_info = "\n" + wait_for_user_response(user_id, "Введите контакты для отклика:")
    send_message(user_id, "Прикрепите картинку.")

    # После получения всех ответов обрабатываем вложение
    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.to_me:
            attachments = vk.messages.getById(message_ids=event.message_id, extended=True, fields='attachments')['items'][0]['attachments']
            if attachments:
                attachment = attachments[0]
                if attachment['type'] == 'photo':
                    photo = attachment['photo']
                    photo_link = event.attachments['attach1']
                    finalText = organization_name + website_link + job_title + job_function + job_requirements + job_conditions + contact_info
                    download_photo(finalText, photo_link, photo, "photo.jpg")
                    send_message(user_id, message="Ваша вакансия на модерации.\nПросим заметить, что для нас приоритетными являются вакансии, напрямую связанные с направлениями подготовки, реализуемые в университете.")
                    break
                else:
                    send_message(user_id, "Найдено вложение другого типа")
                    break
            else:
                send_message(user_id, "Вложений не обнаружено.")
                break


def wait_for_user_response(user_id, prompt_message):
    # Отправляем сообщение с запросом пользователю и ожидаем его ответа
    send_message(user_id, message=prompt_message)
    
    # Получаем ответ от пользователя
    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.user_id == user_id:
            send_message(user_id, "Ответ записан.")
            return event.text.strip()  # Возвращаем текст ответа пользователя



# Авторизация в ВКонтакте
token_file = open('C:\\1\\token1.txt')
vk_session = vk_api.VkApi(token=token_file.read())
token_file.close()
vk = vk_session.get_api()
longpoll = VkLongPoll(vk_session)

token_file_post = open('C:\\1\\token2.txt')
vk_session_post = vk_api.VkApi(token=token_file_post.read())
token_file_post.close()
vk_post = vk_session_post.get_api()

admin_id_file = open('C:\\1\\admin.txt')
id_admin = admin_id_file.read()
admin_id_file.close()

group_id_file = open('C:\\1\\id_group.txt')
id_group = group_id_file.read()
group_id_file.close()


current_dir = os.path.dirname(os.path.abspath(__file__))

stop_flag = threading.Event()

def publication():
    while not stop_flag.is_set():
        time.sleep(2.5)
        publ_post()

def main():
    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW:
            if event.to_me:
                msg =event.text.lower()
                id = event.user_id

                if id == id_admin:
                    adminValid = True
                else:
                    adminValid = False

                if event.text.lower().startswith('!рек')& adminValid:
                    command = msg.split()[1]
                    handle_command(id, command)
                elif (msg =='!id') & adminValid:
                    send_message(id,id)
                elif msg == 'начало':
                    handle_advertisement(id, event)
                elif (msg == 'кон') & adminValid:
                    stop_flag.set()
                    # Ожидаем завершения потока
                    thread_pub.join()
                    print("Поток остановлен")
                    send_message(577763695,"Работа бота остановлена")
                    break

# Создаем поток для функции publication
thread_pub = threading.Thread(target=publication)
# Запускаем поток
thread_pub.start()
main()
